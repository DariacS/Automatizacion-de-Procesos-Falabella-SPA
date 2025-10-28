import os
import subprocess
import sys
import stat
import time
import shutil
import threading
import queue
import logging
import getpass
import webbrowser
import zipfile
import tarfile
import tempfile
from datetime import datetime
from difflib import SequenceMatcher
import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox, ttk, scrolledtext
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# -------------------------------
# Config
# -------------------------------
CONFIG_FILE = "ultima_carpeta.txt"
SUPPORTED_INPUT_EXT = (".xlsx", ".xls", ".xlsm", ".csv", ".json", ".parquet",
                       ".ods", ".txt", ".html", ".htm")
ARCHIVE_EXT = (".zip", ".tar", ".tar.gz", ".tgz", ".gz", ".bz2", ".7z")
LOG_FILENAME = "process.log"
HISTORICO_FILENAME_XLSX = "historico_completo.xlsx"
HISTORICO_FILENAME_CSV = "historico_completo.csv"
NULOS_DETALLE_FILENAME = "detalle_nulos.csv"

# -------------------------------
# Utilidades
# -------------------------------

def timestamp_now(fmt="%Y%m%d_%H%M%S"):
    return datetime.now().strftime(fmt)


def human_now():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def safe_mkdir(path):
    os.makedirs(path, exist_ok=True)
    return path


def calcular_similitud(a, b):
    return SequenceMatcher(None, a.lower(), b.lower()).ratio()


def clasificar_tendencia(porcentaje, filas_actual, filas_base, umbral_filas_min=10):
    """
    Clasifica la tendencia según porcentaje y tamaño de archivo.
    
    - porcentaje: porcentaje de cambio
    - filas_actual: filas del archivo actual
    - filas_base: filas del mes anterior o promedio histórico
    - umbral_filas_min: umbral mínimo de filas para forzar 'Cambio Leve'
    """

    # Si el tamaño del archivo o referencia es menor al umbral, forzar "Cambio Leve"
    try:
        fa = float(filas_actual) if filas_actual is not None else None
        fb = float(filas_base) if filas_base is not None else None
    except Exception:
        fa, fb = filas_actual, filas_base

    if (fa is not None and fa < umbral_filas_min) or (fb is not None and fb < umbral_filas_min):
        return "Cambio Leve"

    # Clasificación según porcentaje
    abs_pct = abs(porcentaje)
    if abs_pct <= 10:
        return "Cambio Leve"
    elif abs_pct <= 30:
        return "Cambio Moderado"
    elif abs_pct <= 50:
        return "Cambio Alto"
    else:
        return "Cambio Crítico"

# -------------------------------
# Función para extraer mes/año desde carpeta (MMYYYY)
# -------------------------------

def extraer_mes_anio(nombre_carpeta):
    if len(nombre_carpeta) == 6 and nombre_carpeta.isdigit():
        mes = nombre_carpeta[:2]
        anio = nombre_carpeta[2:]
        return f"{mes}/{anio}"
    return "Desconocido"

# -------------------------------
# Buscar carpetas de mes dentro de carpeta base (solo carpetas con formato MMYYYY)
# -------------------------------

def buscar_carpetas_mes(carpeta_base):
    carpetas_mes = []
    for entry in os.listdir(carpeta_base):
        ruta = os.path.join(carpeta_base, entry)
        if os.path.isdir(ruta) and len(entry) == 6 and entry.isdigit():
            carpetas_mes.append(entry)
    carpetas_mes.sort()  # orden cronológico ascendente
    return carpetas_mes

# -------------------------------
# Utilidades archivos / comprimidos
# -------------------------------

def is_archive_path(path):
    ln = path.lower()
    return any(ln.endswith(a) for a in ARCHIVE_EXT)


def is_csv_gz(path):
    name = os.path.basename(path).lower()
    return name.endswith(".csv.gz") or name.endswith(".csv.bz2") or name.endswith(".csv.xz")

def rmtree_safe(path):
    """Elimina una carpeta aunque tenga archivos de solo lectura o bloqueados."""
    def onerror(func, path, exc_info):
        # Cambiar permiso a escritura y reintentar
        os.chmod(path, stat.S_IWRITE)
        func(path)

    for _ in range(5):  # reintenta 5 veces
        if os.path.exists(path):
            try:
                shutil.rmtree(path, onerror=onerror)
                return True
            except Exception:
                time.sleep(0.1)
        else:
            return True
    return False

def _sanitize_zipinfo_name(name):
    # Evitar path traversal: eliminar rutas absolutas y ..
    # Mantener solo el basename o la ruta relativa segura
    # Reemplaza backslashes por slashes
    name = name.replace('\\', '/')
    # eliminar prefijo / si existe
    if name.startswith('/'):
        name = name.lstrip('/')
    parts = [p for p in name.split('/') if p not in ('', '..')]
    return '/'.join(parts)

def limpiar_temp_antiguos(temp_root, logger=None):
    """
    Elimina subcarpetas antiguas en temp_root dejando solo la última.
    """
    try:
        if not os.path.exists(temp_root):
            return

        carpetas = [os.path.join(temp_root, d) for d in os.listdir(temp_root) 
                    if os.path.isdir(os.path.join(temp_root, d))]

        if not carpetas:
            return

        # Ordenar por fecha de modificación
        carpetas.sort(key=os.path.getmtime, reverse=True)

        # Mantener solo la primera (más reciente)
        for carpeta in carpetas[1:]:
            try:
                shutil.rmtree(carpeta, ignore_errors=True)
                if logger: logger.info(f"Carpeta temporal eliminada: {carpeta}")
            except Exception as e:
                if logger: logger.warning(f"No se pudo eliminar {carpeta}: {e}")

    except Exception as e:
        if logger: logger.error(f"Error limpiando carpetas temporales: {e}")

def manejar_error_espacio(temp_root, logger=None):
    """
    Maneja la falta de espacio con limpieza de temporales y opciones de continuar.
    """
    root = tk.Tk()
    root.withdraw()

    # Paso 1: Avisar y limpiar temp antiguos
    messagebox.showwarning(
        "Espacio insuficiente",
        "El disco se quedó sin espacio.\n\n"
        "Se eliminarán carpetas temporales antiguas dejando solo la última en uso."
    )
    limpiar_temp_antiguos(os.path.dirname(temp_root), logger)

    # Verificar si ya hay espacio suficiente
    # (aquí puedes agregar tu chequeo real con shutil.disk_usage si quieres)
    espacio = shutil.disk_usage(os.path.dirname(temp_root))
    if espacio.free > 200*1024*1024:  # ejemplo: 200 MB libres mínimo
        return "continuar"

    # Paso 2: Preguntar si desea liberar manualmente
    respuesta = messagebox.askyesno(
        "Liberar espacio",
        "No fue suficiente con eliminar carpetas temporales.\n\n"
        "¿Quieres abrir el explorador de archivos para liberar espacio y luego continuar?"
    )

    if respuesta:
        # Abrir explorador en C:\
        try:
            subprocess.Popen(["explorer", os.path.dirname(temp_root)])
        except:
            pass

        messagebox.showinfo(
            "Continuar",
            "Libera espacio en disco y presiona Aceptar para continuar."
        )
        return "continuar"
    else:
        # Eliminar todo temp y cortar proceso
        try:
            shutil.rmtree(temp_root, ignore_errors=True)
            if logger: logger.info(f"Carpeta {temp_root} eliminada por falta de espacio.")
        except Exception as e:
            if logger: logger.error(f"No se pudo eliminar {temp_root}: {e}")
        return "terminar"

def extraer_archivo(archivo_path, dest_dir):
    """
    Extrae el archivo comprimido en dest_dir.
    Devuelve lista de rutas extraídas (paths completos).
    Soporta zip, tar, gz (single file) y 7z (si py7zr está instalado).
    """
    extraidos = []
    try:
        if zipfile.is_zipfile(archivo_path):
            with zipfile.ZipFile(archivo_path, 'r') as z:
                # sanitizar nombres
                for member in z.infolist():
                    safe_name = _sanitize_zipinfo_name(member.filename)
                    if not safe_name:
                        continue
                    target_path = os.path.join(dest_dir, safe_name)
                    safe_mkdir(os.path.dirname(target_path))
                    if member.is_dir():
                        safe_mkdir(target_path)
                    else:
                        with z.open(member) as src, open(target_path, 'wb') as dst:
                            shutil.copyfileobj(src, dst)
                        extraidos.append(target_path)
        elif tarfile.is_tarfile(archivo_path):
            with tarfile.open(archivo_path, 'r:*') as t:
                for member in t.getmembers():
                    if not member.isreg():
                        continue
                    safe_name = _sanitize_zipinfo_name(member.name)
                    target_path = os.path.join(dest_dir, safe_name)
                    safe_mkdir(os.path.dirname(target_path))
                    f = t.extractfile(member)
                    if f is None:
                        continue
                    with open(target_path, 'wb') as out_f:
                        shutil.copyfileobj(f, out_f)
                    extraidos.append(target_path)
        else:
            name = os.path.basename(archivo_path)
            if name.lower().endswith(('.gz', '.bz2', '.xz')) and not name.lower().endswith('.tar.gz'):
                # Descomprimir archivo único con gzip/bz2/xz
                try:
                    # Intentar shutil.unpack_archive si reconoce
                    shutil.unpack_archive(archivo_path, dest_dir)
                    for root, dirs, files in os.walk(dest_dir):
                        for f in files:
                            extraidos.append(os.path.join(root, f))
                except Exception:
                    # Fallback gzip
                    try:
                        import gzip
                        target_name = os.path.splitext(name)[0]
                        out_path = os.path.join(dest_dir, target_name)
                        with gzip.open(archivo_path, 'rb') as src, open(out_path, 'wb') as dst:
                            shutil.copyfileobj(src, dst)
                        extraidos.append(out_path)
                    except Exception as e:
                        raise RuntimeError(f"No se pudo extraer {archivo_path}: {e}")
            else:
                # intentar py7zr para 7z
                try:
                    import py7zr
                    with py7zr.SevenZipFile(archivo_path, mode='r') as z:
                        z.extractall(path=dest_dir)
                    for root, dirs, files in os.walk(dest_dir):
                        for f in files:
                            extraidos.append(os.path.join(root, f))
                except Exception:
                    raise RuntimeError(f"Formato de archivo comprimido no soportado: {archivo_path}")
    except Exception as e:
        raise RuntimeError(f"Error extrayendo {archivo_path}: {e}")
    return [p for p in extraidos if os.path.exists(p)]

# -------------------------------
# Expandir y listar archivos para procesar (incluye extracción de comprimidos)
# -------------------------------

def expandir_y_listar_archivos_para_procesar(carpeta, temp_root_base=None):
    resultado = []
    if temp_root_base is None:
        temp_root_base = os.path.join(carpeta, ".temp_processing")
    safe_mkdir(temp_root_base)

    for root, dirs, files in os.walk(carpeta):
        parts = os.path.normpath(root).split(os.sep)
        if any(p.lower() == "resumen_generado" for p in parts):
            continue
        for fname in files:
            if fname.startswith("~$"):
                continue
            full = os.path.join(root, fname)
            name_lower = fname.lower()

            if is_archive_path(name_lower):
                ts = timestamp_now()
                tempdir = os.path.join(temp_root_base, f"extract_{ts}_{abs(hash(full))%100000}")
                safe_mkdir(tempdir)
                try:
                    extracted_paths = extraer_archivo(full, tempdir)
                except Exception:
                    resultado.append({"path": full, "original": full, "internal_path": ""})
                    continue

                for p in extracted_paths:
                    bn = os.path.basename(p)
                    if bn.startswith(".") or bn.upper().startswith("__MACOSX"):
                        continue
                    if os.path.isfile(p):
                        if is_archive_path(p):
                            # re-extraer anidado recursivamente
                            resultado.extend(expandir_y_listar_archivos_para_procesar(os.path.dirname(p), temp_root_base=temp_root_base))
                        else:
                            resultado.append({"path": p, "original": full, "internal_path": os.path.relpath(p, tempdir)})
                continue

                        # Añadir TODOS los archivos (no solo los soportados por pandas).
            # El worker decidirá si puede leer el contenido o sólo registrar nombre+size.
            resultado.append({"path": full, "original": full, "internal_path": ""})

    return resultado

# -------------------------------
# Leer archivo genérico
# -------------------------------

def leer_archivo_generico(ruta):
    ext = os.path.splitext(ruta)[1].lower()
    name = os.path.basename(ruta).lower()
    size_mb = os.path.getsize(ruta) / (1024 * 1024)

    # CSV / TXT
    if name.endswith('.csv') or name.endswith('.csv.gz') or name.endswith('.csv.bz2') or name.endswith('.csv.xz') or ext == '.txt':
        if size_mb > 10:  # usar modo streaming para grandes
            with open(ruta, "r", encoding="utf-8", errors="ignore") as f:
                filas = sum(1 for _ in f) - 1  # restar cabecera
            return pd.DataFrame({"__filas__": [filas]})
        else:
            for encoding in ["utf-8-sig", "utf-8", "latin1"]:
                for sep in [",", ";", "\t"]:
                    try:
                        df = pd.read_csv(ruta, encoding=encoding, sep=sep, low_memory=False)
                        if df.shape[1] > 1 or sep == ",":
                            return df
                    except Exception:
                        continue
            raise RuntimeError(f"No se pudo leer CSV/TXT {ruta}")

    # Excel antiguo (.xls)
    if ext == '.xls':
        try:
            import xlrd
            book = xlrd.open_workbook(ruta, on_demand=True)
            sheet = book.sheet_by_index(0)
            filas = sheet.nrows - 1
            return pd.DataFrame({"__filas__": [filas]})
        except Exception as e:
            raise RuntimeError(f"No se pudo leer .xls: {e}")


    # Excel moderno (.xlsx, .xlsm) o ODS
    if ext in ('.xlsx', '.xlsm', '.ods'):
        if size_mb > 10:  # usar openpyxl en modo streaming
            from openpyxl import load_workbook
            wb = load_workbook(ruta, read_only=True, data_only=True)
            ws = wb.active
            filas = ws.max_row - 1
            return pd.DataFrame({"__filas__": [filas]})
        else:
            try:
                df = pd.read_excel(ruta, engine=None)
                return df
            except Exception:
                df = pd.read_excel(ruta)
                return df

    # JSON
    if ext == '.json':
        if size_mb > 10:
            with open(ruta, "r", encoding="utf-8", errors="ignore") as f:
                filas = sum(1 for _ in f)
            return pd.DataFrame({"__filas__": [filas]})
        else:
            try:
                df = pd.read_json(ruta)
                if isinstance(df, dict):
                    df = pd.DataFrame([df])
                return df
            except Exception:
                df = pd.read_json(ruta, lines=True)
                return df

    # Parquet
    if ext == '.parquet':
        try:
            df = pd.read_parquet(ruta)
            return df
        except Exception as e:
            raise RuntimeError(f"Parquet requiere pyarrow/fastparquet: {e}")

    # HTML
    if ext in ('.html', '.htm'):
        tablas = pd.read_html(ruta)
        if tablas:
            return tablas[0]
        else:
            return pd.DataFrame()

    # Otros logs u otros
    if ext in ('.log', '.conf'):
        if size_mb > 10:
            with open(ruta, "r", encoding="utf-8", errors="ignore") as f:
                filas = sum(1 for _ in f)
            return pd.DataFrame({"__filas__": [filas]})
        else:
            try:
                return pd.read_table(ruta)
            except Exception:
                return pd.DataFrame()

    raise RuntimeError(f"Formato no soportado: {ruta}")

# -------------------------------
# Obtener tamaño archivo en KB
# -------------------------------
    
def obtener_tamano_kb(ruta):
    """Devuelve tamaño en KB como float."""
    try:
        size_bytes = os.path.getsize(ruta)
        return size_bytes / 1024  # KB
    except:
        return None


def obtener_tamano_legible(ruta):
    try:
        size_bytes = os.path.getsize(ruta)
        # Definir unidades
        if size_bytes < 1024:
            return f"{size_bytes} B"
        elif size_bytes < 1024**2:
            return f"{size_bytes/1024:.2f} KB"
        elif size_bytes < 1024**3:
            return f"{size_bytes/(1024**2):.2f} MB"
        else:
            return f"{size_bytes/(1024**3):.2f} GB"
    except:
        return None

# -------------------------------
# Formatear archivo Excel generado
# -------------------------------

def formatear_excel(path_excel):
    try:
        wb = load_workbook(path_excel)
        ws = wb.active

        # Estilos
        header_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")  # Amarillo suave
        header_font = Font(bold=True, color="000000")
        center_align = Alignment(horizontal="center", vertical="center")
        right_align = Alignment(horizontal="right", vertical="center")

        tendencia_color_map = {
            "Cambio Leve": "92D050",
            "Cambio Moderado": "FFC000",
            "Cambio Alto": "FF0000",
            "Cambio Crítico": "800000",
            "": None
        }

        ws.freeze_panes = "A2"

        for col_idx, col in enumerate(ws.columns, start=1):
            max_length = 0
            col_letter = get_column_letter(col_idx)

            cell_header = ws[f"{col_letter}1"]
            cell_header.fill = header_fill
            cell_header.font = header_font
            cell_header.alignment = center_align

            for cell in col:
                if cell.value is not None:
                    longitud = len(str(cell.value))
                    if longitud > max_length:
                        max_length = longitud

            adjusted_width = max(10, max_length + 2)
            ws.column_dimensions[col_letter].width = adjusted_width

        encabezados = {ws.cell(row=1, column=col_idx).value: col_idx for col_idx in range(1, ws.max_column+1)}

        for col_name in ["Tamaño KB", "Tamaño archivo mes anterior"]:
            if col_name in encabezados:
                col = encabezados[col_name]
                for row in range(2, ws.max_row + 1):
                    c = ws.cell(row=row, column=col)
                    if isinstance(c.value, (int, float)):
                        size_bytes = c.value * 1024  # convertir KB a bytes
                        if size_bytes < 1024:
                            c.value = size_bytes
                            c.number_format = '0.00 "B"'
                        elif size_bytes < 1024**2:
                            c.value = size_bytes / 1024
                            c.number_format = '0.00 "KB"'
                        elif size_bytes < 1024**3:
                            c.value = size_bytes / (1024**2)
                            c.number_format = '0.00 "MB"'
                        else:
                            c.value = size_bytes / (1024**3)
                            c.number_format = '0.00 "GB"'
                        c.alignment = right_align


        for col_name in ["Tendencia % vs mes anterior", "Tendencia histórica %"]:
            if col_name in encabezados:
                col = encabezados[col_name]
                for row in range(2, ws.max_row + 1):
                    c = ws.cell(row=row, column=col)
                    if isinstance(c.value, (int, float)):
                        c.number_format = '0.00"%"'
                        c.alignment = right_align

        if "Clasificación tendencia mes anterior" in encabezados:
            col = encabezados["Clasificación tendencia mes anterior"]
            for row in range(2, ws.max_row + 1):
                c = ws.cell(row=row, column=col)
                color_hex = tendencia_color_map.get(c.value, None)
                if color_hex:
                    c.fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")

        if "Clasificación tendencia histórica" in encabezados:
            col = encabezados["Clasificación tendencia histórica"]
            for row in range(2, ws.max_row + 1):
                c = ws.cell(row=row, column=col)
                color_hex = tendencia_color_map.get(c.value, None)
                if color_hex:
                    c.fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")

        # Formatear tendencia de peso en %
        for col_name in ["Tendencia % peso vs mes anterior", "Tendencia histórica % peso"]:
            if col_name in encabezados:
                col = encabezados[col_name]
                for row in range(2, ws.max_row + 1):
                    c = ws.cell(row=row, column=col)
                    if isinstance(c.value, (int, float)):
                        c.number_format = '0.00"%"'
                        c.alignment = right_align

        # Colorear clasificación de tendencia de peso
        for col_name in ["Clasificación tendencia peso mes anterior", "Clasificación tendencia peso histórica"]:
            if col_name in encabezados:
                col = encabezados[col_name]
                for row in range(2, ws.max_row + 1):
                    c = ws.cell(row=row, column=col)
                    color_hex = tendencia_color_map.get(c.value, None)
                    if color_hex:
                        c.fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")

        for col_name in ["Cantidad filas", "Cantidad filas mes anterior"]:
            if col_name in encabezados:
                col = encabezados[col_name]
                for row in range(2, ws.max_row + 1):
                    c = ws.cell(row=row, column=col)
                    if isinstance(c.value, (int, float)):
                        c.number_format = "#,##0"
                        c.alignment = right_align

        wb.save(path_excel)
    except Exception as e:
        print(f"Error formateando archivo Excel {path_excel}: {e}")

# -------------------------------
# Mover a escritorio
# -------------------------------
def mover_a_backup_escritorio(ruta_path, carpeta_base, tipo="Archivo o carpeta"):
    try:
        escritorio = os.path.join(os.path.expanduser("~"), "Desktop")

        # Tomamos el nombre de la carpeta_base como raíz del backup
        nombre_base = os.path.basename(carpeta_base.rstrip(os.sep))

        # Agregar fecha y hora legible (YYYYMMDD_HHMMSS)
        fecha_hora = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_root = os.path.join(escritorio, "CMF_Backup", f"{nombre_base}_{fecha_hora}")

        safe_mkdir(backup_root)

        # Obtener la ruta relativa (ejemplo: "012025/resumen/archivo.xlsx")
        ruta_relativa = os.path.relpath(ruta_path, carpeta_base)

        # Crear la misma estructura dentro del backup
        destino = os.path.join(backup_root, ruta_relativa)

        # Crear subcarpetas necesarias
        os.makedirs(os.path.dirname(destino), exist_ok=True)

        # Mover archivo o carpeta
        shutil.move(ruta_path, destino)

        return destino
    except Exception as e:
        print(f"No se pudo mover {tipo} a backup: {e}")
        return None

# -------------------------------
# Preguntar y borrar duplicados
# -------------------------------

def preguntar_y_borrar_duplicados_tk(carpeta_base):
    archivo_historico_xlsx = os.path.join(carpeta_base, HISTORICO_FILENAME_XLSX)
    archivo_historico_csv = os.path.join(carpeta_base, HISTORICO_FILENAME_CSV)

    carpetas_mes = buscar_carpetas_mes(carpeta_base)
    root = tk.Tk()
    root.withdraw()

    # Archivos históricos
    for ruta_hist in [archivo_historico_xlsx, archivo_historico_csv]:
        if os.path.isfile(ruta_hist):
            respuesta = messagebox.askyesno(
                "Archivo histórico existente",
                f"El archivo '{os.path.basename(ruta_hist)}' ya existe en la carpeta base.\n"
                "¿Desea moverlo a carpeta de respaldo en escritorio para evitar duplicados?"
            )
            if respuesta:
                mover_a_backup_escritorio(ruta_hist, carpeta_base, tipo="Archivo")
            else:
                root.destroy()
                return False

    # Carpetas de resumen por mes
    for mes in carpetas_mes:
        carpeta_resumen = os.path.join(carpeta_base, mes, "resumen_generado")
        if os.path.isdir(carpeta_resumen):
            respuesta = messagebox.askyesno(
                "Carpeta resumen existente",
                f"La carpeta 'resumen_generado' existe en {mes}.\n"
                "¿Desea moverla a carpeta de respaldo en escritorio para evitar duplicados?"
            )
            if respuesta:
                mover_a_backup_escritorio(carpeta_resumen, carpeta_base, tipo="Carpeta")
            else:
                root.destroy()
                return False

    root.destroy()
    return True

# -------------------------------
# Contar filas válidas (no vacías) en DataFrame
# -------------------------------

def contar_filas_validas(df):
    # Caso especial: si viene del modo rápido (archivos grandes)
    if "__filas__" in df.columns and df.shape[1] == 1:
        return int(df["__filas__"].iloc[0])

    # Caso normal: limpiar y contar
    try:
        # Eliminar filas totalmente vacías
        df = df.dropna(how='all')
        return len(df)
    except Exception:
        return 0

# -------------------------------
# Proceso completo (worker)
# -------------------------------

def worker_procesar(carpeta_base, q, cancel_event, logger, total_carpetas=None,umbral_filas_min=10):
    logger.info("Inicio de procesamiento en carpeta base: %s", carpeta_base)
    q.put(("info", f"Iniciando procesamiento: {human_now()}"))

    ruta_resumen_generado = os.path.join(carpeta_base, "resumen_generado")
    ruta_historico_completo = os.path.join(carpeta_base, "historico_completo")

    if not preguntar_y_borrar_duplicados_tk(carpeta_base):
        logger.info("Proceso cancelado por el usuario.")
        q.put(("info", "Proceso cancelado por el usuario."))
        return

    carpetas_mes = buscar_carpetas_mes(carpeta_base)
    if not carpetas_mes:
        q.put(("info", "No se encontraron carpetas de meses con formato MMYYYY"))
        q.put(("done", {}))
        return

    logger.info(f"Carpetas de meses encontradas: {carpetas_mes}")
    # Carpeta temporal base en el Escritorio
    usuario = getpass.getuser()
    escritorio = os.path.join("C:\\Users", usuario, "Desktop")
    temp_root_base = os.path.join(escritorio, "cmf_temp")
    safe_mkdir(temp_root_base)


    # Cargar histórico existente
    ruta_historico_xlsx = os.path.join(carpeta_base, HISTORICO_FILENAME_XLSX)
    ruta_historico_csv = os.path.join(carpeta_base, HISTORICO_FILENAME_CSV)
    df_historico = pd.DataFrame()
    if os.path.exists(ruta_historico_xlsx):
        try:
            df_historico = pd.read_excel(ruta_historico_xlsx)
            logger.info(f"Histórico cargado: {ruta_historico_xlsx}, filas: {len(df_historico)}")
        except Exception as e:
            logger.warning(f"No se pudo cargar histórico: {e}")
            df_historico = pd.DataFrame()

    resultados_mensuales = {}
    total_meses = len(carpetas_mes)

    for idx_mes, mes_carpeta in enumerate(carpetas_mes, start=1):
        print(f"[DEBUG] Inicio procesamiento mes {mes_carpeta} ({idx_mes}/{total_meses})")
        logger.info(f"[DEBUG] Inicio procesamiento mes {mes_carpeta} ({idx_mes}/{total_meses})")
        if cancel_event.is_set():
            logger.info("Proceso cancelado por usuario.")
            q.put(("info", "Proceso cancelado por usuario."))
            break

        porcentaje = int((idx_mes - 1) / total_meses * 100)
        q.put(("progress", porcentaje))
        q.put(("info", f"Procesando mes: {mes_carpeta}"))

        ruta_mes = os.path.join(carpeta_base, mes_carpeta)
        temp_root = os.path.join(temp_root_base, f"{mes_carpeta}_{timestamp_now()}")
        safe_mkdir(temp_root)

        items = []
        while True:
            try:
                items = expandir_y_listar_archivos_para_procesar(ruta_mes, temp_root_base=temp_root)
                break  # si tuvo éxito, salir del bucle
            except OSError as e:
                if hasattr(e, "winerror") and e.winerror == 112:  # Espacio insuficiente
                    logger.error(f"Espacio insuficiente en disco durante la expansión de archivos del mes {mes_carpeta}.")
                    q.put(("info", f"Espacio insuficiente en disco en mes {mes_carpeta}. Intentando liberar espacio..."))

                    decision = manejar_error_espacio(temp_root, logger)

                    if decision == "continuar":
                        logger.info(f"Reintentando expansión de archivos mes {mes_carpeta} después de liberar espacio...")
                        try:
                            if os.path.isdir(temp_root):
                                shutil.rmtree(temp_root)
                        except Exception as e2:
                            logger.warning(f"No se pudo eliminar temp_root: {e2}")
                        safe_mkdir(temp_root)
                        continue  # vuelve a intentar
                    else:
                        q.put(("info", f"Proceso detenido por falta de espacio en disco en mes {mes_carpeta}."))
                        return
                else:
                    raise

        # Eliminar duplicados por nombre
        nombres_vistos = set()
        archivos = []
        for item in items:
            nombre_a = os.path.basename(item['path'])
            if nombre_a not in nombres_vistos:
                nombres_vistos.add(nombre_a)
                archivos.append(item)

        if not archivos:
            q.put(("info", f"No hay archivos en carpeta mes {mes_carpeta}"))
            try:
                if os.path.isdir(temp_root):
                    shutil.rmtree(temp_root)
            except Exception as e:
                logger.warning(f"No se pudo limpiar temp dir {temp_root}: {e}")
            continue

        datos_mes = []
        correlativo = 1
        detalle_nulos = []
        total_archivos = len(archivos)

        for idx_archivo, item in enumerate(archivos, start=1):
            if cancel_event.is_set():
                logger.info("Cancelación solicitada por usuario durante mes.")
                break

            archivo_path = item['path']

            # Determinar nombre final a mostrar
            if item.get('internal_path'):
                nombre_archivo = item['internal_path'].replace("\\", "/")
            else:
                ruta_relativa = os.path.relpath(item['path'], ruta_mes)
                nombre_archivo = ruta_relativa.replace("\\", "/")

            size_kb = obtener_tamano_kb(item['original']) if os.path.exists(item['original']) else obtener_tamano_kb(item['path'])

            try:
                mensaje = f"Procesando archivo ({idx_archivo}/{total_archivos}) del mes {mes_carpeta}: {nombre_archivo}"
                print(mensaje)  # Para consola
                q.put(("info", mensaje))  # Para la GUI
                df = leer_archivo_generico(archivo_path)
                read_error = None
            except Exception as e:
                logger.info(f"No se pudo leer como tabla {nombre_archivo}: {e}")
                df = None
                read_error = e

            try:
                df = leer_archivo_generico(archivo_path)
                read_error = None
            except Exception as e:
                logger.info(f"No se pudo leer como tabla {nombre_archivo}: {e}")
                df = None
                read_error = e

            filas_validas = contar_filas_validas(df) if isinstance(df, pd.DataFrame) else 0

            # --- Detalle de nulos ---
            if isinstance(df, pd.DataFrame) and not df.empty:
                mask_nulos = df.isnull()
                for col in df.select_dtypes(include=['object']).columns:
                    mask_nulos[col] = mask_nulos[col] | (df[col].astype(str).str.strip() == "")
                filas_con_nulos = df[mask_nulos.any(axis=1)]
                for i, (fila_idx, fila) in enumerate(filas_con_nulos.iterrows(), start=2):
                    fila_mask = mask_nulos.loc[fila_idx]
                    cols_afectadas = [col for col in df.columns if fila_mask[col].any()]
                    detalle_nulos.append({
                        "Archivo": nombre_archivo,
                        "Ruta": item['path'],
                        "Fila": i,
                        "Columnas_Afectadas": ", ".join(str(c) for c in cols_afectadas if c is not None)
                    })

            datos_mes.append({
                "Correlativo": correlativo,
                "Nombre archivo": nombre_archivo,
                "Tamaño KB": size_kb,
                "Cantidad filas": filas_validas,
                "Carpeta Mes": mes_carpeta
            })
            correlativo += 1

            porcentaje_carpeta = int(idx_archivo / total_archivos * 100)
            q.put(("progress_carpeta", porcentaje_carpeta))

        # --- Tendencias vs mes anterior e histórico ---
        mes_anterior = carpetas_mes[idx_mes - 2] if idx_mes > 1 else None
        datos_tendencia = []

        df_mes = pd.DataFrame(datos_mes)

        if not df_mes.empty:
            datos_tendencia = []

        for _, row in df_mes.iterrows():
            nombre_actual = row["Nombre archivo"]
            size_actual = row["Tamaño KB"]
            filas_actual = row["Cantidad filas"]

            # --- Primer mes: sin comparación ---
            if mes_anterior is None:
                datos_tendencia.append({
                    "Correlativo": row["Correlativo"],
                    "Nombre archivo": nombre_actual,
                    "Tamaño KB": size_actual,
                    "Cantidad filas": filas_actual,
                    "Carpeta Mes": mes_carpeta,
                    # Agregar columnas de peso aunque sea "sin comparación"
                    "Tendencia % vs mes anterior": None,
                    "Clasificación tendencia mes anterior": "Sin comparación",
                    "Tendencia histórica %": None,
                    "Clasificación tendencia histórica": "Sin comparación",
                    "Tendencia % peso vs mes anterior": None,
                    "Clasificación tendencia peso mes anterior": "Sin comparación",
                    "Tendencia histórica % peso": None,
                    "Clasificación tendencia peso histórica": "Sin comparación",
                })
                continue

            # --- Meses siguientes: comparación con mes anterior ---
            fila_mes_ant = None
            if not df_historico.empty:
                df_mes_anterior = df_historico[df_historico["Carpeta Mes"] == mes_anterior].copy()
                if not df_mes_anterior.empty:
                    exact = df_mes_anterior[df_mes_anterior["Nombre archivo"] == nombre_actual]
                    if not exact.empty:
                        fila_mes_ant = exact.iloc[-1]
                    else:
                        df_mes_anterior["_sim"] = df_mes_anterior["Nombre archivo"].apply(lambda x: calcular_similitud(x, nombre_actual))
                        df_mes_anterior = df_mes_anterior.sort_values("_sim", ascending=False)
                        if not df_mes_anterior.empty and df_mes_anterior.iloc[0]["_sim"] >= 0.8:
                            fila_mes_ant = df_mes_anterior.iloc[0]

            # --- Tendencia filas ---
            tendencia_pct = None
            tendencia_clas = "Sin comparación"
            if fila_mes_ant is not None:
                try:
                    filas_ant_val = float(fila_mes_ant.get("Cantidad filas", None))
                except Exception:
                    filas_ant_val = None

                if filas_ant_val is not None and filas_ant_val != 0:
                    try:
                        if filas_ant_val < 10 and filas_actual < 10:
                            tendencia_pct = None
                            tendencia_clas = "Cambio Leve"
                        else:
                            tendencia_pct = ((float(filas_actual) - filas_ant_val) / filas_ant_val) * 100
                            tendencia_pct = round(tendencia_pct, 2)
                            tendencia_clas = clasificar_tendencia(
                                tendencia_pct,
                                filas_actual,
                                filas_ant_val,
                                umbral_filas_min
                            )
                    except Exception:
                        tendencia_pct = None
                        tendencia_clas = "Sin comparación"

            # --- Tendencia peso ---
            tendencia_peso_pct = None
            tendencia_peso_clas = "Sin comparación"
            if fila_mes_ant is not None:
                try:
                    peso_ant_val = float(fila_mes_ant.get("Tamaño KB", None))
                except Exception:
                    peso_ant_val = None

                if peso_ant_val is not None and peso_ant_val != 0:
                    try:
                        if peso_ant_val < 10 and size_actual < 10:
                            tendencia_peso_pct = None
                            tendencia_peso_clas = "Cambio Leve"
                        else:
                            tendencia_peso_pct = ((size_actual - peso_ant_val) / peso_ant_val) * 100
                            tendencia_peso_pct = round(tendencia_peso_pct, 2)
                            tendencia_peso_clas = clasificar_tendencia(
                                tendencia_peso_pct,
                                size_actual,
                                peso_ant_val,
                                umbral_filas_min
                            )
                    except Exception:
                        tendencia_peso_pct = None
                        tendencia_peso_clas = "Sin comparación"

            # --- Tendencia histórica (filas) ---
            tendencia_hist_pct = None
            tendencia_hist_clas = "Sin comparación"
            if not df_historico.empty:
                mask_hist = (df_historico["Carpeta Mes"] != mes_carpeta) & (
                    (df_historico["Nombre archivo"] == nombre_actual) |
                    (df_historico["Nombre archivo"].apply(lambda x: calcular_similitud(x, nombre_actual) >= 0.8))
                )
                df_similares = df_historico[mask_hist]

                if not df_similares.empty:
                    try:
                        promedio_hist = float(df_similares["Cantidad filas"].mean())
                    except Exception:
                        promedio_hist = None

                    if promedio_hist and promedio_hist != 0:
                        try:
                            tendencia_hist_pct = ((float(filas_actual) - promedio_hist) / promedio_hist) * 100
                            tendencia_hist_pct = round(tendencia_hist_pct, 2)
                            tendencia_hist_clas = clasificar_tendencia(
                                tendencia_hist_pct,
                                filas_actual,
                                promedio_hist,
                                umbral_filas_min
                            )
                        except Exception:
                            tendencia_hist_pct = None
                            tendencia_hist_clas = "Sin comparación"

            # --- Tendencia histórica (peso) ---
            tendencia_peso_hist_pct = None
            tendencia_peso_hist_clas = "Sin comparación"
            if not df_historico.empty:
                mask_hist_peso = (df_historico["Carpeta Mes"] != mes_carpeta) & (
                    (df_historico["Nombre archivo"] == nombre_actual) |
                    (df_historico["Nombre archivo"].apply(lambda x: calcular_similitud(x, nombre_actual) >= 0.8))
                )
                df_similares_peso = df_historico[mask_hist_peso]

                if not df_similares_peso.empty:
                    try:
                        promedio_peso_hist = float(df_similares_peso["Tamaño KB"].mean())
                    except Exception:
                        promedio_peso_hist = None

                    if promedio_peso_hist and promedio_peso_hist != 0:
                        try:
                            tendencia_peso_hist_pct = ((float(size_actual) - promedio_peso_hist) / promedio_peso_hist) * 100
                            tendencia_peso_hist_pct = round(tendencia_peso_hist_pct, 2)
                            tendencia_peso_hist_clas = clasificar_tendencia(
                                tendencia_peso_hist_pct,
                                size_actual,
                                promedio_peso_hist,
                                umbral_filas_min
                            )
                        except Exception:
                            tendencia_peso_hist_pct = None
                            tendencia_peso_hist_clas = "Sin comparación"

            # --- Agregar al resultado ---
            datos_tendencia.append({
                "Correlativo": row["Correlativo"],
                "Nombre archivo": nombre_actual,
                "Tamaño KB": size_actual,
                "Cantidad filas": filas_actual,
                "Nombre archivo mes anterior": fila_mes_ant["Nombre archivo"] if fila_mes_ant is not None else "",
                "Tamaño archivo mes anterior": fila_mes_ant["Tamaño KB"] if fila_mes_ant is not None else None,
                "Cantidad filas mes anterior": fila_mes_ant["Cantidad filas"] if fila_mes_ant is not None else None,
                "Tendencia % vs mes anterior": tendencia_pct,
                "Clasificación tendencia mes anterior": tendencia_clas,
                "Tendencia histórica %": tendencia_hist_pct,
                "Clasificación tendencia histórica": tendencia_hist_clas,
                "Tendencia % peso vs mes anterior": tendencia_peso_pct,
                "Clasificación tendencia peso mes anterior": tendencia_peso_clas,
                "Tendencia histórica % peso": tendencia_peso_hist_pct,
                "Clasificación tendencia peso histórica": tendencia_peso_hist_clas,
                "Carpeta Mes": mes_carpeta
            })

        df_mes = pd.DataFrame(datos_tendencia)
        print(f"[DEBUG] Fin procesamiento mes {mes_carpeta} ({idx_mes}/{total_meses})")
        logger.info(f"[DEBUG] Fin procesamiento mes {mes_carpeta} ({idx_mes}/{total_meses})")

        # --- Guardar archivos ---
        carpeta_salida_mes = os.path.join(ruta_mes, "resumen_generado")
        safe_mkdir(carpeta_salida_mes)
        nombre_base = f"resumen_{mes_carpeta}"
        ruta_excel_mes = os.path.join(carpeta_salida_mes, f"{nombre_base}.xlsx")
        ruta_csv_mes = os.path.join(carpeta_salida_mes, f"{nombre_base}.csv")

        try:
            columnas_excluir = [
                                    "Tendencia histórica %",
                                    "Clasificación tendencia histórica",
                                    "Tendencia histórica % peso",
                                    "Clasificación tendencia peso histórica",
                                ]
            columnas_finales = [c for c in df_mes.columns if c not in columnas_excluir]
            df_mes_sin_hist = df_mes[columnas_finales]

            df_mes_sin_hist.to_excel(ruta_excel_mes, index=False)
            formatear_excel(ruta_excel_mes)
            df_mes_sin_hist.to_csv(ruta_csv_mes, index=False)

            logger.info(f"Guardado resumen mensual en {ruta_excel_mes} y {ruta_csv_mes}")
            q.put(("info", f"Guardado resumen mensual mes {mes_carpeta}"))
        except Exception as e:
            logger.error(f"Error guardando resumen mensual mes {mes_carpeta}: {e}")

        if detalle_nulos:
            ruta_detalle_nulos = os.path.join(carpeta_salida_mes, NULOS_DETALLE_FILENAME)
            try:
                df_nulos = pd.DataFrame(detalle_nulos)
                df_nulos.to_csv(ruta_detalle_nulos, index=False)
                logger.info(f"Guardado detalle nulos en {ruta_detalle_nulos}")
                q.put(("info", f"Detalle de nulos guardado mes {mes_carpeta}"))
            except Exception as e:
                logger.error(f"Error guardando detalle nulos: {e}")

        if not df_mes.empty:
            resultados_mensuales[mes_carpeta] = df_mes
            if "Carpeta Mes" in df_historico.columns:
                df_historico = df_historico[df_historico["Carpeta Mes"] != mes_carpeta]

            # --- 🔹 Solo guardar columnas necesarias ---
            columnas_historico = [
                "Correlativo",
                "Nombre archivo",
                "Tamaño KB",
                "Cantidad filas",
                "Carpeta Mes",
                "Tendencia histórica %",
                "Clasificación tendencia histórica",
                "Tendencia histórica % peso",
                "Clasificación tendencia peso histórica",
            ]
            cols_presentes = [c for c in columnas_historico if c in df_mes.columns]
            df_historico_mes = df_mes[cols_presentes]

            df_historico = pd.concat([df_historico, df_historico_mes], ignore_index=True)


        porcentaje = int(idx_mes / total_meses * 100)
        q.put(("progress", porcentaje))
        q.put(("progress_carpeta", 0))

        # limpiar temporales del mes
        try:
            if os.path.isdir(temp_root):
                shutil.rmtree(temp_root)
        except Exception as e:
            logger.warning(f"No se pudo limpiar temp dir {temp_root}: {e}")

    # --- Guardar histórico completo ---
    try:
        ruta_historico_xlsx_out = os.path.join(carpeta_base, HISTORICO_FILENAME_XLSX)
        ruta_historico_csv_out = os.path.join(carpeta_base, HISTORICO_FILENAME_CSV)
        df_historico.to_excel(ruta_historico_xlsx_out, index=False)
        formatear_excel(ruta_historico_xlsx_out)
        df_historico.to_csv(ruta_historico_csv_out, index=False)
        logger.info(f"Guardado histórico completo en {ruta_historico_xlsx_out} y {ruta_historico_csv_out}")
        q.put(("info", f"Histórico completo actualizado"))
    except Exception as e:
        logger.error(f"Error guardando histórico completo: {e}")

    # Borrar carpeta temporal completa al final
    if os.path.isdir(temp_root_base):
        if rmtree_safe(temp_root_base):
            logger.info(f"Carpeta temporal {temp_root_base} eliminada correctamente")
        else:
            logger.warning(f"No se pudo eliminar la carpeta temporal {temp_root_base} después de varios intentos")



    q.put(("done", {"total_meses": len(resultados_mensuales)}))
    logger.info("Proceso finalizado.")

# -------------------------------
# GUI principal
# -------------------------------

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Analizador Tendencia Archivos - Local")
        self.geometry("1000x500")  
        self.resizable(True, True)

        self.last_folder = None
        self.worker_thread = None
        self.cancel_event = threading.Event()
        self.queue = queue.Queue()

        self._setup_widgets()
        self._load_last_folder()

        # Logger básico
        logging.basicConfig(filename=LOG_FILENAME, level=logging.INFO,
                            format='%(asctime)s - %(levelname)s - %(message)s')
        self.logger = logging.getLogger()

        # Iniciar ciclo de proceso de cola
        self._process_queue()

    def _setup_widgets(self):
        pad = 7
        # ====== Estilo general ======
        style = ttk.Style(self)
        self.configure(bg="#f0f2f5")  # Fondo general suave
        style.configure("TLabel", background="#f0f2f5", font=("Segoe UI", 11))
        style.configure("TEntry", font=("Segoe UI", 11))
        style.configure("TButton", font=("Segoe UI", 10, "bold"), padding=6)
        style.configure("Progress.Horizontal.TProgressbar", thickness=18, troughcolor="#d0d0d0", background="#4caf50")
        style.configure("Blue.Horizontal.TProgressbar", thickness=18, troughcolor="#d0d0d0", background="#2196f3")

        # ====== Frame superior: selección carpeta + botones ======
        frame_top = tk.Frame(self, bg="#ffffff", relief="flat", bd=0)
        frame_top.pack(fill=tk.X, padx=pad, pady=(pad, 5))
        frame_top.config(highlightbackground="#cccccc", highlightthickness=1)

        lbl = ttk.Label(frame_top, text="📁", font=("Segoe UI", 12, "bold"))
        lbl.pack(side=tk.LEFT, padx=(0,5), pady=5)

        # Label para mostrar carpeta seleccionada, truncando si es muy larga
        self.folder_display_var = tk.StringVar(value="Selecciona una carpeta...")
        self.lbl_folder_display = tk.Label(frame_top, textvariable=self.folder_display_var,
                                           font=("Segoe UI", 11), bg="#f0f0f0", anchor="w",
                                           relief="sunken", bd=1, width=50)
        self.lbl_folder_display.pack(side=tk.LEFT, padx=(0,5), pady=5, fill=tk.X, expand=True)

        def actualizar_ruta_label(ruta):
            # Trunca la ruta si es demasiado larga
            max_len = 50
            if len(ruta) > max_len:
                parts = ruta.split(os.sep)
                truncated = "…" + os.sep.join(parts[-(max_len//2):])
                self.folder_display_var.set(truncated)
            else:
                self.folder_display_var.set(ruta)

        # Botones con estilo
        def style_button(btn, bg_color, hover_color):
            btn.config(relief="flat", bd=0, highlightthickness=0, cursor="hand2", bg=bg_color, fg="#ffffff")
            btn.bind("<Enter>", lambda e: btn.config(bg=hover_color))
            btn.bind("<Leave>", lambda e: btn.config(bg=bg_color))

        btn_browse = tk.Button(frame_top, text="Seleccionar carpeta", command=self.seleccionar_carpeta)
        style_button(btn_browse, "#4caf50", "#45a049")
        btn_browse.pack(side=tk.LEFT, padx=(0,5), pady=5)

        btn_start = tk.Button(frame_top, text="Iniciar procesamiento", command=self.iniciar_proceso)
        style_button(btn_start, "#2196f3", "#1976d2")
        btn_start.pack(side=tk.LEFT, padx=(0,5), pady=5)

        btn_cancel = tk.Button(frame_top, text="Cancelar proceso", command=self.cancelar_proceso)
        style_button(btn_cancel, "#f44336", "#d32f2f")
        btn_cancel.pack(side=tk.LEFT, padx=(0,5), pady=5)

        # ====== Barra de progreso general ======
        progress_frame = tk.Frame(self, bg="#f0f2f5")
        progress_frame.pack(fill=tk.X, padx=pad, pady=(5, 5))

        self.progress_var = tk.DoubleVar(value=0)
        self.progress_bar = ttk.Progressbar(progress_frame, orient="horizontal", mode="determinate",
                                            maximum=100, variable=self.progress_var,
                                            style="Progress.Horizontal.TProgressbar")
        self.progress_bar.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(5,0), pady=5)

        self.lbl_progress = ttk.Label(progress_frame, text="0%", font=("Segoe UI", 10, "bold"))
        self.lbl_progress.pack(side=tk.LEFT, padx=(5,5))

        # ====== Barra de progreso de carpeta actual ======
        progress_frame_carpeta = tk.Frame(self, bg="#f0f2f5")
        progress_frame_carpeta.pack(fill=tk.X, padx=pad, pady=(0,5))

        self.progress_var_carpeta = tk.DoubleVar(value=0)
        self.progress_bar_carpeta = ttk.Progressbar(progress_frame_carpeta, orient="horizontal", mode="determinate",
                                                    maximum=100, variable=self.progress_var_carpeta,
                                                    style="Blue.Horizontal.TProgressbar")
        self.progress_bar_carpeta.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(5,0), pady=5)

        self.lbl_progress_carpeta = ttk.Label(progress_frame_carpeta, text="0%", font=("Segoe UI", 10, "bold"))
        self.lbl_progress_carpeta.pack(side=tk.LEFT, padx=(5,5))

        # ====== Parámetro: Umbral mínimo de filas ======
        frame_umbral = tk.Frame(self, bg="#f0f2f5")
        frame_umbral.pack(fill=tk.X, padx=pad, pady=(5, 5))

        lbl_umbral = ttk.Label(frame_umbral, text="Umbral mínimo de filas:")
        lbl_umbral.pack(side=tk.LEFT, padx=(5,5))

        self.umbral_var = tk.IntVar(value=10)  # valor por defecto
        spin_umbral = tk.Spinbox(frame_umbral, from_=1, to=100,
                                 textvariable=self.umbral_var, width=5)
        spin_umbral.pack(side=tk.LEFT, padx=(0,10))

        lbl_explicacion = ttk.Label(
            frame_umbral,
            text="Si el archivo y su referencia de mes anterior tienen menos filas que este valor, el cambio se clasifica como 'Cambio Leve'.",
            font=("Segoe UI", 9)
        )
        lbl_explicacion.pack(side=tk.LEFT, padx=(5,5))


        # ====== Log y salida ======
        text_frame = tk.Frame(self, bg="#2b2b3a", relief="flat")
        text_frame.pack(fill=tk.BOTH, expand=True, padx=pad, pady=(0,pad))
        text_frame.config(highlightbackground="#444", highlightthickness=1)

        self.text_log = scrolledtext.ScrolledText(text_frame, height=25, wrap=tk.WORD,
                                                  bg="#1e1e2f", fg="#ffffff", insertbackground="#ffffff",
                                                  font=("Consolas", 10), relief="flat")
        self.text_log.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)

        # Guardar referencia para actualizar carpeta seleccionada
        self.actualizar_ruta_label = actualizar_ruta_label

    def _load_last_folder(self):
        if os.path.exists(CONFIG_FILE):
            try:
                with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                    carpeta = f.read().strip()
                    if os.path.isdir(carpeta):
                        self.last_folder = carpeta
                        self.actualizar_ruta_label(carpeta)
            except Exception as e:
                self.logger.warning(f"No se pudo leer configuración: {e}")

    def _save_last_folder(self, carpeta):
        try:
            with open(CONFIG_FILE, "w", encoding="utf-8") as f:
                f.write(carpeta)
        except Exception as e:
            self.logger.warning(f"No se pudo guardar configuración: {e}")

    def seleccionar_carpeta(self):
        carpeta = filedialog.askdirectory()
        if carpeta:
            self.last_folder = carpeta
            self.actualizar_ruta_label(carpeta)
            self._save_last_folder(carpeta)

    def iniciar_proceso(self):
        carpeta = self.last_folder
        if not carpeta or not os.path.isdir(carpeta):
            messagebox.showerror("Error", "Por favor selecciona una carpeta base válida.")
            return

        if not preguntar_y_borrar_duplicados_tk(carpeta):
            self._log("Proceso cancelado por el usuario antes de iniciar.")
            return

        if self.worker_thread and self.worker_thread.is_alive():
            messagebox.showinfo("Información", "Proceso ya está en ejecución.")
            return

        self.text_log.configure(state="normal")
        self.text_log.delete(1.0, tk.END)
        self.text_log.configure(state="disabled")

        self.progress_var.set(0)
        self.lbl_progress.config(text="0%")
        self.progress_var_carpeta.set(0)
        self.lbl_progress_carpeta.config(text="0%")

        self.cancel_event.clear()

        # Total carpetas para calcular avance
        carpetas = [c for c in os.listdir(carpeta) if os.path.isdir(os.path.join(carpeta, c))]
        total_carpetas = len(carpetas)

        self.worker_thread = threading.Thread(
            target=worker_procesar,
            args=(carpeta, self.queue, self.cancel_event, self.logger, total_carpetas),
            daemon=True
        )
        self.worker_thread.start()

        self.worker_thread = threading.Thread(
        target=worker_procesar,
        args=(carpeta, self.queue, self.cancel_event, self.logger, total_carpetas, self.umbral_var.get()),
        daemon=True)


    def cancelar_proceso(self):
        if self.worker_thread and self.worker_thread.is_alive():
            self.cancel_event.set()
            self.logger.info("Cancelación solicitada por usuario.")
            self._log("Cancelando proceso...")

    def _log(self, texto):
        self.text_log.configure(state="normal")
        self.text_log.insert(tk.END, texto + "\n")
        self.text_log.see(tk.END)
        self.text_log.configure(state="disabled")

    def _process_queue(self):
        try:
            while True:
                msg = self.queue.get_nowait()
                if msg[0] == "info":
                    self._log(msg[1])
                elif msg[0] == "progress":
                    valor = msg[1]
                    self.progress_var.set(valor)
                    self.lbl_progress.config(text=f"{valor}%")
                elif msg[0] == "progress_carpeta":
                    valor = msg[1]
                    self.progress_var_carpeta.set(valor)
                    self.lbl_progress_carpeta.config(text=f"{valor}%")
                elif msg[0] == "done":
                    total = msg[1].get("total_meses", 0)
                    self._log(f"Proceso finalizado. Total meses procesados: {total}")
                    self.progress_var.set(100)
                    self.lbl_progress.config(text="100%")
                    self.progress_var_carpeta.set(0)
                    self.lbl_progress_carpeta.config(text="0%")
                self.queue.task_done()
        except queue.Empty:
            pass
        self.after(200, self._process_queue)

# -------------------------------
# Programa principal
# -------------------------------

if __name__ == "__main__":
    app = App()
    app.mainloop()
