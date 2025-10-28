import os
import subprocess
import sys
import stat
import time
import shutil
import threading
import queue
import logging
import getpass
from venv import logger
import webbrowser
import zipfile
import tarfile
import tempfile
from datetime import datetime
from difflib import SequenceMatcher
import pandas as pd
import tkinter as tk
from tkinter import filedialog, messagebox, ttk, scrolledtext
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

# -------------------------------
# Config
# -------------------------------
CONFIG_FILE = "ultima_carpeta.txt"
SUPPORTED_INPUT_EXT = (".xlsx", ".xls", ".xlsm", ".csv", ".json", ".parquet",
                       ".ods", ".txt", ".html", ".htm")
ARCHIVE_EXT = (".zip", ".tar", ".tar.gz", ".tgz", ".gz", ".bz2", ".7z")
LOG_FILENAME = "process.log"
HISTORICO_FILENAME_XLSX = "resumen_completo.xlsx"
HISTORICO_FILENAME_CSV = "resumen_completo.csv"
NULOS_DETALLE_FILENAME = "detalle_nulos.csv"
HISTORICO_XLSX = "historico_completo.xlsx"
HISTORICO_CSV = "historico_completo.csv"
# -------------------------------
# Utilidades
# -------------------------------

def timestamp_now(fmt="%Y%m%d_%H%M%S"):
    return datetime.now().strftime(fmt)

def human_now():
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")

def safe_mkdir(path):
    os.makedirs(path, exist_ok=True)
    return path

# -------------------------------
# Función para extraer mes/año desde carpeta (MMYYYY)
# -------------------------------
def extraer_mes_anio(nombre_carpeta):
    if len(nombre_carpeta) == 6 and nombre_carpeta.isdigit():
        mes = nombre_carpeta[:2]
        anio = nombre_carpeta[2:]
        return f"{mes}/{anio}"
    return "Desconocido"

# -------------------------------
# Buscar carpetas de mes dentro de carpeta base (solo carpetas con formato MMYYYY)
# -------------------------------
def buscar_carpetas_mes(carpeta_base):
    carpetas_mes = []
    for entry in os.listdir(carpeta_base):
        ruta = os.path.join(carpeta_base, entry)
        if os.path.isdir(ruta) and len(entry) == 6 and entry.isdigit():
            carpetas_mes.append(entry)
    carpetas_mes.sort()  # orden cronológico ascendente
    return carpetas_mes
# -------------------------------
# Utilidades archivos / comprimidos
# -------------------------------
def is_archive_path(path):
    ln = path.lower()
    return any(ln.endswith(a) for a in ARCHIVE_EXT)

def is_csv_gz(path):
    name = os.path.basename(path).lower()
    return name.endswith(".csv.gz") or name.endswith(".csv.bz2") or name.endswith(".csv.xz")

def rmtree_safe(path):
    """Elimina una carpeta aunque tenga archivos de solo lectura o bloqueados."""
    def onerror(func, path, exc_info):
        os.chmod(path, stat.S_IWRITE)
        func(path)
    for _ in range(5):
        if os.path.exists(path):
            try:
                shutil.rmtree(path, onerror=onerror)
                return True
            except Exception:
                time.sleep(0.1)
        else:
            return True
    return False

def _sanitize_zipinfo_name(name):
    name = name.replace('\\', '/')
    if name.startswith('/'):
        name = name.lstrip('/')
    parts = [p for p in name.split('/') if p not in ('', '..')]
    return '/'.join(parts)

def limpiar_temp_antiguos(temp_root, logger=None):
    try:
        if not os.path.exists(temp_root):
            return
        carpetas = [os.path.join(temp_root, d) for d in os.listdir(temp_root) 
                    if os.path.isdir(os.path.join(temp_root, d))]
        if not carpetas:
            return
        carpetas.sort(key=os.path.getmtime, reverse=True)
        for carpeta in carpetas[1:]:
            try:
                shutil.rmtree(carpeta, ignore_errors=True)
                if logger: logger.info(f"Carpeta temporal eliminada: {carpeta}")
            except Exception as e:
                if logger: logger.warning(f"No se pudo eliminar {carpeta}: {e}")
    except Exception as e:
        if logger: logger.error(f"Error limpiando carpetas temporales: {e}")

def manejar_error_espacio(temp_root, logger=None):
    root = tk.Tk()
    root.withdraw()
    messagebox.showwarning(
        "Espacio insuficiente",
        "El disco se quedó sin espacio.\n\n"
        "Se eliminarán carpetas temporales antiguas dejando solo la última en uso."
    )
    limpiar_temp_antiguos(os.path.dirname(temp_root), logger)
    espacio = shutil.disk_usage(os.path.dirname(temp_root))
    if espacio.free > 200*1024*1024:
        return "continuar"
    respuesta = messagebox.askyesno(
        "Liberar espacio",
        "No fue suficiente con eliminar carpetas temporales.\n\n"
        "¿Quieres abrir el explorador de archivos para liberar espacio y luego continuar?"
    )
    if respuesta:
        try:
            subprocess.Popen(["explorer", os.path.dirname(temp_root)])
        except:
            pass
        messagebox.showinfo(
            "Continuar",
            "Libera espacio en disco y presiona Aceptar para continuar."
        )
        return "continuar"
    else:
        try:
            shutil.rmtree(temp_root, ignore_errors=True)
            if logger: logger.info(f"Carpeta {temp_root} eliminada por falta de espacio.")
        except Exception as e:
            if logger: logger.error(f"No se pudo eliminar {temp_root}: {e}")
        return "terminar"

def extraer_archivo(archivo_path, dest_dir):
    extraidos = []
    try:
        if zipfile.is_zipfile(archivo_path):
            with zipfile.ZipFile(archivo_path, 'r') as z:
                for member in z.infolist():
                    safe_name = _sanitize_zipinfo_name(member.filename)
                    if not safe_name:
                        continue
                    target_path = os.path.join(dest_dir, safe_name)
                    safe_mkdir(os.path.dirname(target_path))
                    if member.is_dir():
                        safe_mkdir(target_path)
                    else:
                        with z.open(member) as src, open(target_path, 'wb') as dst:
                            shutil.copyfileobj(src, dst)
                        extraidos.append(target_path)
        elif tarfile.is_tarfile(archivo_path):
            with tarfile.open(archivo_path, 'r:*') as t:
                for member in t.getmembers():
                    if not member.isreg():
                        continue
                    safe_name = _sanitize_zipinfo_name(member.name)
                    target_path = os.path.join(dest_dir, safe_name)
                    safe_mkdir(os.path.dirname(target_path))
                    f = t.extractfile(member)
                    if f is None:
                        continue
                    with open(target_path, 'wb') as out_f:
                        shutil.copyfileobj(f, out_f)
                    extraidos.append(target_path)
        else:
            name = os.path.basename(archivo_path)
            if name.lower().endswith(('.gz', '.bz2', '.xz')) and not name.lower().endswith('.tar.gz'):
                try:
                    shutil.unpack_archive(archivo_path, dest_dir)
                    for root, dirs, files in os.walk(dest_dir):
                        for f in files:
                            extraidos.append(os.path.join(root, f))
                except Exception:
                    try:
                        import gzip
                        target_name = os.path.splitext(name)[0]
                        out_path = os.path.join(dest_dir, target_name)
                        with gzip.open(archivo_path, 'rb') as src, open(out_path, 'wb') as dst:
                            shutil.copyfileobj(src, dst)
                        extraidos.append(out_path)
                    except Exception as e:
                        raise RuntimeError(f"No se pudo extraer {archivo_path}: {e}")
            else:
                try:
                    import py7zr
                    with py7zr.SevenZipFile(archivo_path, mode='r') as z:
                        z.extractall(path=dest_dir)
                    for root, dirs, files in os.walk(dest_dir):
                        for f in files:
                            extraidos.append(os.path.join(root, f))
                except Exception:
                    raise RuntimeError(f"Formato de archivo comprimido no soportado: {archivo_path}")
    except Exception as e:
        raise RuntimeError(f"Error extrayendo {archivo_path}: {e}")
    return [p for p in extraidos if os.path.exists(p)]

# -------------------------------
# Expandir y listar archivos para procesar
# -------------------------------
def expandir_y_listar_archivos_para_procesar(carpeta, temp_root_base=None):
    resultado = []
    if temp_root_base is None:
        temp_root_base = os.path.join(carpeta, ".temp_processing")
    safe_mkdir(temp_root_base)
    for root, dirs, files in os.walk(carpeta):
        parts = os.path.normpath(root).split(os.sep)
        if any(p.lower() == "resumen_generado" for p in parts):
            continue
        for fname in files:
            if fname.startswith("~$"):
                continue
            full = os.path.join(root, fname)
            name_lower = fname.lower()
            if is_archive_path(name_lower):
                ts = timestamp_now()
                tempdir = os.path.join(temp_root_base, f"extract_{ts}_{abs(hash(full))%100000}")
                safe_mkdir(tempdir)
                try:
                    extracted_paths = extraer_archivo(full, tempdir)
                except Exception:
                    resultado.append({"path": full, "original": full, "internal_path": ""})
                    continue
                for p in extracted_paths:
                    bn = os.path.basename(p)
                    if bn.startswith(".") or bn.upper().startswith("__MACOSX"):
                        continue
                    if os.path.isfile(p):
                        if is_archive_path(p):
                            resultado.extend(expandir_y_listar_archivos_para_procesar(os.path.dirname(p), temp_root_base=temp_root_base))
                        else:
                            resultado.append({"path": p, "original": full, "internal_path": os.path.relpath(p, tempdir)})
                continue
            resultado.append({"path": full, "original": full, "internal_path": ""})
    return resultado

# -------------------------------
# Leer archivo genérico
# -------------------------------
def leer_archivo_generico(ruta):
    ext = os.path.splitext(ruta)[1].lower()
    name = os.path.basename(ruta).lower()
    size_mb = os.path.getsize(ruta) / (1024 * 1024)
    if name.endswith('.csv') or name.endswith('.csv.gz') or name.endswith('.csv.bz2') or name.endswith('.csv.xz') or ext == '.txt':
        if size_mb > 10:
            with open(ruta, "r", encoding="utf-8", errors="ignore") as f:
                filas = sum(1 for _ in f) - 1
            return pd.DataFrame({"__filas__": [filas]})
        else:
            for encoding in ["utf-8-sig", "utf-8", "latin1"]:
                for sep in [",", ";", "\t"]:
                    try:
                        df = pd.read_csv(ruta, encoding=encoding, sep=sep, low_memory=False)
                        if df.shape[1] > 1 or sep == ",":
                            return df
                    except Exception:
                        continue
            raise RuntimeError(f"No se pudo leer CSV/TXT {ruta}")
    if ext == '.xls':
        try:
            import xlrd
            book = xlrd.open_workbook(ruta, on_demand=True)
            sheet = book.sheet_by_index(0)
            filas = sheet.nrows - 1
            return pd.DataFrame({"__filas__": [filas]})
        except Exception as e:
            raise RuntimeError(f"No se pudo leer .xls: {e}")
    if ext in ('.xlsx', '.xlsm', '.ods'):
        if size_mb > 10:
            wb = load_workbook(ruta, read_only=True, data_only=True)
            ws = wb.active
            filas = ws.max_row - 1
            return pd.DataFrame({"__filas__": [filas]})
        else:
            try:
                df = pd.read_excel(ruta, engine=None)
                return df
            except Exception:
                df = pd.read_excel(ruta)
                return df
    if ext == '.json':
        if size_mb > 10:
            with open(ruta, "r", encoding="utf-8", errors="ignore") as f:
                filas = sum(1 for _ in f)
            return pd.DataFrame({"__filas__": [filas]})
        else:
            try:
                df = pd.read_json(ruta)
                if isinstance(df, dict):
                    df = pd.DataFrame([df])
                return df
            except Exception:
                df = pd.read_json(ruta, lines=True)
                return df
    if ext == '.parquet':
        try:
            df = pd.read_parquet(ruta)
            return df
        except Exception as e:
            raise RuntimeError(f"Parquet requiere pyarrow/fastparquet: {e}")
    if ext in ('.html', '.htm'):
        tablas = pd.read_html(ruta)
        if tablas:
            return tablas[0]
        else:
            return pd.DataFrame()
    if ext in ('.log', '.conf'):
        if size_mb > 10:
            with open(ruta, "r", encoding="utf-8", errors="ignore") as f:
                filas = sum(1 for _ in f)
            return pd.DataFrame({"__filas__": [filas]})
        else:
            try:
                return pd.read_table(ruta)
            except Exception:
                return pd.DataFrame()
    raise RuntimeError(f"Formato no soportado: {ruta}")

# -------------------------------
# Obtener tamaño archivo
# -------------------------------
def obtener_tamano_kb(ruta):
    try:
        size_bytes = os.path.getsize(ruta)
        return size_bytes / 1024
    except:
        return None

def obtener_tamano_legible(ruta):
    try:
        size_bytes = os.path.getsize(ruta)
        if size_bytes < 1024:
            return f"{size_bytes} B"
        elif size_bytes < 1024**2:
            return f"{size_bytes/1024:.2f} KB"
        elif size_bytes < 1024**3:
            return f"{size_bytes/(1024**2):.2f} MB"
        else:
            return f"{size_bytes/(1024**3):.2f} GB"
    except:
        return None

# -------------------------------
# Formatear archivo Excel generado
# -------------------------------
def formatear_excel(path_excel):
    try:
        wb = load_workbook(path_excel)
        ws = wb.active
        header_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
        header_font = Font(bold=True, color="000000")
        center_align = Alignment(horizontal="center", vertical="center")
        right_align = Alignment(horizontal="right", vertical="center")
        ws.freeze_panes = "A2"
        for col_idx, col in enumerate(ws.columns, start=1):
            max_length = 0
            col_letter = get_column_letter(col_idx)
            cell_header = ws[f"{col_letter}1"]
            cell_header.fill = header_fill
            cell_header.font = header_font
            cell_header.alignment = center_align
            for cell in col:
                if cell.value is not None:
                    longitud = len(str(cell.value))
                    if longitud > max_length:
                        max_length = longitud
            adjusted_width = max(10, max_length + 2)
            ws.column_dimensions[col_letter].width = adjusted_width
        encabezados = {ws.cell(row=1, column=col_idx).value: col_idx for col_idx in range(1, ws.max_column+1)}
        for col_name in ["Tamaño KB", "Tamaño archivo mes anterior"]:
            if col_name in encabezados:
                col = encabezados[col_name]
                for row in range(2, ws.max_row + 1):
                    c = ws.cell(row=row, column=col)
                    if isinstance(c.value, (int, float)):
                        size_bytes = c.value * 1024
                        if size_bytes < 1024:
                            c.value = size_bytes
                            c.number_format = '0.00 "B"'
                        elif size_bytes < 1024**2:
                            c.value = size_bytes / 1024
                            c.number_format = '0.00 "KB"'
                        elif size_bytes < 1024**3:
                            c.value = size_bytes / (1024**2)
                            c.number_format = '0.00 "MB"'
                        else:
                            c.value = size_bytes / (1024**3)
                            c.number_format = '0.00 "GB"'
                        c.alignment = right_align
        for col_name in ["Cantidad filas"]:
            if col_name in encabezados:
                col = encabezados[col_name]
                for row in range(2, ws.max_row + 1):
                    c = ws.cell(row=row, column=col)
                    if isinstance(c.value, (int, float)):
                        c.number_format = "#,##0"
                        c.alignment = right_align
        wb.save(path_excel)

    except Exception as e:
        print(f"Error formateando archivo Excel {path_excel}: {e}")

# -------------------------------
# Mover a escritorio
# -------------------------------
def mover_a_backup_escritorio(ruta_path, carpeta_base, tipo="Archivo o carpeta"):
    try:
        escritorio = os.path.join(os.path.expanduser("~"), "Desktop")
        nombre_base = os.path.basename(carpeta_base.rstrip(os.sep))
        fecha_hora = datetime.now().strftime("%Y%m%d_%H%M%S")
        backup_root = os.path.join(escritorio, "CMF_Backup", f"{nombre_base}_{fecha_hora}")
        safe_mkdir(backup_root)
        ruta_relativa = os.path.relpath(ruta_path, carpeta_base)
        destino = os.path.join(backup_root, ruta_relativa)
        os.makedirs(os.path.dirname(destino), exist_ok=True)
        shutil.move(ruta_path, destino)
        return destino
    except Exception as e:
        print(f"No se pudo mover {tipo} a backup: {e}")
        return None

# -------------------------------
# Preguntar y borrar duplicados
# -------------------------------
def preguntar_y_borrar_duplicados_tk(carpeta_base, logger=None):
    # Ignorar archivos históricos
    archivo_historico_xlsx = os.path.join(carpeta_base, HISTORICO_FILENAME_XLSX)
    archivo_historico_csv = os.path.join(carpeta_base, HISTORICO_FILENAME_CSV)
    for ruta_hist in [archivo_historico_xlsx, archivo_historico_csv]:
        if os.path.isfile(ruta_hist):
            if logger:
                logger.info(f"Ignorando archivo histórico existente: {ruta_hist}")

    # Ignorar carpetas resumen_generado
    carpetas_mes = buscar_carpetas_mes(carpeta_base)
    for mes in carpetas_mes:
        carpeta_resumen = os.path.join(carpeta_base, mes, "resumen_generado")
        if os.path.isdir(carpeta_resumen):
            if logger:
                logger.info(f"Ignorando carpeta existente 'resumen_generado' en {mes}")

    return True

# -------------------------------
# Procesamiento de carpeta completa
# -------------------------------
def procesar_carpeta_base(carpeta_base, logger=None, cancel_event=None, progress_callback=None, progress_carpeta_callback=None):
    import time  # solo si usas simulación de progreso
    if cancel_event is None:
        from threading import Event
        cancel_event = Event()

    if not preguntar_y_borrar_duplicados_tk(carpeta_base):
        if logger: logger.info("Proceso cancelado por usuario.")
        return

    carpetas_mes = buscar_carpetas_mes(carpeta_base)
    historico_data = []
    nulos_detalle = []

    total_carpetas = len(carpetas_mes)
    # Carpeta temporal en el escritorio
    escritorio = os.path.join(os.path.expanduser("~"), "Desktop")
    temp_root = os.path.join(escritorio, ".temp_processing")
    safe_mkdir(temp_root)


    for idx, mes in enumerate(carpetas_mes, start=1):
        if cancel_event.is_set():
            if logger: logger.warning("Proceso cancelado por usuario.")
            return

        ruta_mes = os.path.join(carpeta_base, mes)
        archivos_para_procesar = expandir_y_listar_archivos_para_procesar(ruta_mes, temp_root_base=temp_root)

        for file_idx, item in enumerate(archivos_para_procesar, start=1):
            if cancel_event.is_set():
                if logger: logger.warning("Proceso cancelado por usuario.")
                return

            path = item['path']
            original = item['original']
            internal_path = item['internal_path']

            try:
                df = leer_archivo_generico(path)
                filas = len(df) if '__filas__' not in df.columns else df['__filas__'][0]
                size_kb = obtener_tamano_kb(path)
                historico_data.append({
                    "Mes": extraer_mes_anio(mes),
                    "Archivo": os.path.basename(path),
                    "Cantidad filas": filas,
                    "Tamaño KB": round(size_kb, 2) if size_kb else None,
                    "Ruta completa": path
                })
                if logger: logger.info(f"Procesado: {path} ({filas} filas)")
            except Exception as e:
                nulos_detalle.append({
                    "Mes": extraer_mes_anio(mes),
                    "Archivo": os.path.basename(path),
                    "Ruta completa": path,
                    "Error": str(e)
                })
                if logger: logger.warning(f"Error leyendo {path}: {e}")

            # Actualizar progreso de carpeta
            if progress_carpeta_callback:
                progress_carpeta_callback(file_idx / len(archivos_para_procesar) * 100)

        # Actualizar progreso total
        if progress_callback:
            progress_callback(idx / total_carpetas * 100)

    # Guardar resultados finales (igual que antes)
    if historico_data:
        df_hist = pd.DataFrame(historico_data)
        path_xlsx = os.path.join(carpeta_base, HISTORICO_FILENAME_XLSX)
        path_csv = os.path.join(carpeta_base, HISTORICO_FILENAME_CSV)
        df_hist.to_excel(path_xlsx, index=False)
        df_hist.to_csv(path_csv, index=False)
        formatear_excel(path_xlsx)
        if logger: logger.info(f"Resumen guardado en {path_xlsx} y {path_csv}")

    if nulos_detalle:
        df_nulos = pd.DataFrame(nulos_detalle)
        path_nulos = os.path.join(carpeta_base, NULOS_DETALLE_FILENAME)
        df_nulos.to_csv(path_nulos, index=False)
        if logger: logger.info(f"Detalle de archivos nulos/errores guardado en {path_nulos}")

    limpiar_temp_antiguos(temp_root, logger)
    if logger: logger.info("Procesamiento completo.")


# -------------------------------<
# Configurar Logger
# -------------------------------
def configurar_logger(log_file=None):
    logger = logging.getLogger("ProcesadorCMF")
    logger.setLevel(logging.DEBUG)
    formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
    if log_file:
        fh = logging.FileHandler(log_file, encoding='utf-8')
        fh.setLevel(logging.DEBUG)
        fh.setFormatter(formatter)
        logger.addHandler(fh)
    ch = logging.StreamHandler()
    ch.setLevel(logging.INFO)
    ch.setFormatter(formatter)
    logger.addHandler(ch)
    return logger


# -------------------------------
# Interfaz Tkinter para seleccionar carpeta y ejecutar
# -------------------------------
class TextHandler(logging.Handler):
    """Log handler que envia mensajes a un ScrolledText."""
    def __init__(self, text_widget):
        logging.Handler.__init__(self)
        self.text_widget = text_widget
        self.text_widget.configure(state='disabled')

    def emit(self, record):
        msg = self.format(record)
        def append():
            self.text_widget.configure(state='normal')
            self.text_widget.insert(tk.END, msg + '\n')
            self.text_widget.yview(tk.END)
            self.text_widget.configure(state='disabled')
        self.text_widget.after(0, append)

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Listador Archivos")
        self.geometry("1000x500")
        self.resizable(True, True)

        self.last_folder = None
        self.worker_thread = None
        self.cancel_event = threading.Event()
        self.queue = queue.Queue()

        self._setup_widgets()
        self._load_last_folder()
        self._setup_logger()

    def _setup_logger(self):
        self.logger = logging.getLogger("ProcesadorCMF")
        self.logger.setLevel(logging.DEBUG)
        formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')

        # Log en archivo
        fh = logging.FileHandler(LOG_FILENAME, encoding='utf-8')
        fh.setLevel(logging.DEBUG)
        fh.setFormatter(formatter)
        self.logger.addHandler(fh)

        # Log en ScrolledText
        th = TextHandler(self.text_log)
        th.setLevel(logging.INFO)
        th.setFormatter(formatter)
        self.logger.addHandler(th)

    def _setup_widgets(self):
        pad = 7
        style = ttk.Style(self)
        self.configure(bg="#f0f2f5")
        style.configure("TLabel", background="#f0f2f5", font=("Segoe UI", 11))
        style.configure("TButton", font=("Segoe UI", 10, "bold"), padding=6)
        style.configure("Progress.Horizontal.TProgressbar", thickness=18, troughcolor="#d0d0d0", background="#4caf50")
        style.configure("Blue.Horizontal.TProgressbar", thickness=18, troughcolor="#d0d0d0", background="#2196f3")

        # Frame superior: selección carpeta + botones
        frame_top = tk.Frame(self, bg="#ffffff", relief="flat", bd=0)
        frame_top.pack(fill=tk.X, padx=pad, pady=(pad,5))
        frame_top.config(highlightbackground="#cccccc", highlightthickness=1)

        self.folder_display_var = tk.StringVar(value="Selecciona una carpeta...")
        self.lbl_folder_display = tk.Label(frame_top, textvariable=self.folder_display_var,
                                            font=("Segoe UI", 11), bg="#f0f0f0", anchor="w",
                                            relief="sunken", bd=1, width=50)
        self.lbl_folder_display.pack(side=tk.LEFT, padx=(0,5), pady=5, fill=tk.X, expand=True)

        btn_browse = tk.Button(frame_top, text="Seleccionar carpeta", command=self.seleccionar_carpeta)
        btn_browse.pack(side=tk.LEFT, padx=(0,5), pady=5)

        btn_start = tk.Button(frame_top, text="Iniciar procesamiento", command=self.iniciar_proceso)
        btn_start.pack(side=tk.LEFT, padx=(0,5), pady=5)

        btn_cancel = tk.Button(frame_top, text="Cancelar proceso", command=self.cancelar_proceso)
        btn_cancel.pack(side=tk.LEFT, padx=(0,5), pady=5)

        # ====== Barra de progreso general ======
        progress_frame = tk.Frame(self, bg="#f0f2f5")
        progress_frame.pack(fill=tk.X, padx=pad, pady=(5, 5))

        self.progress_var = tk.DoubleVar(value=0)
        self.progress_bar = ttk.Progressbar(progress_frame, orient="horizontal", mode="determinate",
                                            maximum=100, variable=self.progress_var,
                                            style="Progress.Horizontal.TProgressbar")
        self.progress_bar.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(5,0), pady=5)

        self.lbl_progress = ttk.Label(progress_frame, text="0%", font=("Segoe UI", 10, "bold"))
        self.lbl_progress.pack(side=tk.LEFT, padx=(5,5))

        # ====== Barra de progreso de carpeta actual ======
        progress_frame_carpeta = tk.Frame(self, bg="#f0f2f5")
        progress_frame_carpeta.pack(fill=tk.X, padx=pad, pady=(0,5))

        self.progress_var_carpeta = tk.DoubleVar(value=0)
        self.progress_bar_carpeta = ttk.Progressbar(progress_frame_carpeta, orient="horizontal", mode="determinate",
                                                    maximum=100, variable=self.progress_var_carpeta,
                                                    style="Blue.Horizontal.TProgressbar")
        self.progress_bar_carpeta.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(5,0), pady=5)

        self.lbl_progress_carpeta = ttk.Label(progress_frame_carpeta, text="0%", font=("Segoe UI", 10, "bold"))
        self.lbl_progress_carpeta.pack(side=tk.LEFT, padx=(5,5))

        # Log en ScrolledText
        self.text_log = scrolledtext.ScrolledText(self, height=25, wrap=tk.WORD,
                                                    bg="#1e1e2f", fg="#ffffff", insertbackground="#ffffff",
                                                    font=("Consolas", 10), relief="flat")
        self.text_log.pack(fill=tk.BOTH, expand=True, padx=pad, pady=(0,pad))

    def _load_last_folder(self):
        if os.path.exists(CONFIG_FILE):
            try:
                with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                    carpeta = f.read().strip()
                    if os.path.isdir(carpeta):
                        self.last_folder = carpeta
                        self.folder_display_var.set(carpeta)
            except:
                pass

    def seleccionar_carpeta(self):
        carpeta = filedialog.askdirectory()
        if carpeta:
            self.last_folder = carpeta
            self.folder_display_var.set(carpeta)

    def iniciar_proceso(self):
        if not self.last_folder or not os.path.isdir(self.last_folder):
            messagebox.showerror("Error", "Selecciona primero una carpeta válida.")
            return
        if self.worker_thread and self.worker_thread.is_alive():
            messagebox.showwarning("Proceso en curso", "Ya hay un proceso en ejecución.")
            return
        self.cancel_event.clear()
        self.worker_thread = threading.Thread(target=self._proceso_hilo)
        self.worker_thread.start()

    def cancelar_proceso(self):
        if self.worker_thread and self.worker_thread.is_alive():
            self.cancel_event.set()
            self.logger.info("Se solicitó cancelar el proceso...")

    def _proceso_hilo(self):
        try:
            self.logger.info(f"Iniciando procesamiento en carpeta: {self.last_folder}")

            # Callback para actualizar barras de progreso
            def progreso_total(valor):
                self.progress_var.set(valor)
                self.text_log.update_idletasks()

            def progreso_carpeta(valor):
                self.progress_var_carpeta.set(valor)
                self.text_log.update_idletasks()

            procesar_carpeta_base(
                self.last_folder,
                logger=self.logger,
                cancel_event=self.cancel_event,
                progress_callback=progreso_total,
                progress_carpeta_callback=progreso_carpeta
            )

        except Exception as e:
            self.logger.error(f"Error en proceso: {e}")

if __name__ == "__main__":
    app = App()
    app.mainloop()